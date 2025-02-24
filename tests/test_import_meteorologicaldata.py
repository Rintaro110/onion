import pandas as pd
import openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill

def import_meteorological_month_data(file_path, start_year=None, end_year=None, verbose=False):
    """ 
    月ごとに気象データを取得し、年度ごとの集計を行う関数
    """
    print("---------------------------------------------------")
    print("月別気象データを取得しています...")
    
    # ファイルを開く
    wb = openpyxl.load_workbook(file_path)
    print(f"ファイルを開いた: {file_path}")
    sheet = wb.active

    # 4行目でのラベル取得
    labels = [cell.value for cell in sheet[4][1:]]  # 4行目のB列以降のラベルを取得

    # 6行目で除外する列のチェック
    excluded_labels = ['現象なし情報', '品質情報', '均質番号']
    columns_to_include = []
    
    for idx, cell in enumerate(sheet[6][1:], start=2):  # 6行目のB列以降をチェック
        if not any(excluded_label in str(cell.value) for excluded_label in excluded_labels):
            columns_to_include.append(idx)  # 列のインデックスを保存

    # フィルタリングされたラベルを保持するリスト
    filtered_labels = [labels[idx-2] for idx in columns_to_include]

    # データを[ラベル][年度][月][データ]の形で取得する辞書
    meteorological_data = defaultdict(lambda: defaultdict(dict))
    years_collected = set()  # 取得した年度を記録
    none_count = 0  # 欠損データカウント

    for row in range(7, sheet.max_row + 1):  # 7行目からデータが始まる
        year_month = sheet.cell(row=row, column=1).value
        if not year_month:
            continue

        # 年月から年と月を分離して、会計年度を判定
        year, month = map(int, year_month.split('/'))
        fiscal_year = year if month <= 8 else year + 1

        # 年度フィルタリング（指定された範囲内のみ取得）
        if start_year and fiscal_year < start_year:
            continue
        if end_year and fiscal_year > end_year:
            continue

        years_collected.add(fiscal_year)  # 取得した年度を記録

        for label in filtered_labels:
            if fiscal_year not in meteorological_data[label]:
                meteorological_data[label][fiscal_year] = {}

        # 各ラベルのデータを取得
        for col_idx, label in zip(columns_to_include, filtered_labels):
            if month >= 12 or month <= 8:
                if month not in meteorological_data[label][fiscal_year]:
                    value = sheet.cell(row=row, column=col_idx).value

                    # 空白セル（None）の処理：スキップし、スキップ情報を表示
                    if value is None:
                        none_count += 1
                        """ if verbose:
                            print(f"データスキップ: {fiscal_year}年 {month}月 ({label})") """
                        continue  # このデータをスキップして次の処理へ

                    meteorological_data[label][fiscal_year][month] = value

    # 取得したデータの年度範囲を表示
    if years_collected:
        min_year = min(years_collected)
        max_year = max(years_collected)
        print(f"取得されたデータの年度範囲: {min_year}年 ～ {max_year}年")
    else:
        print("Warning: 指定された範囲に有効なデータがありませんでした。")

    if verbose:
        print(f"labels: {filtered_labels}")
        print(f"総データ数: {sum(len(y) for y in meteorological_data.values())}")
        print(f"欠損データ (None) の数: {none_count}")

    print("気象データ取得完了。")
    print("---------------------------------------------------")

    # データを返す
    return meteorological_data

def import_meteorological_syun_data(file_path, start_year=None, end_year=None, verbose=False):
    """ 
    旬ごとに気象データを取得し、年度ごとの集計を行う関数
    """

    print("旬別気象データを取得しています...")
    print("---------------------------------------------------")

    # ファイルを開く
    wb = openpyxl.load_workbook(file_path)
    print(f"ファイルを開いた: {file_path}")
    sheet = wb.active

    # 4行目でのラベル取得
    labels = [cell.value for cell in sheet[4][1:]]  # 4行目のB列以降のラベルを取得

    # 6行目で除外する列のチェック
    excluded_labels = ['現象なし情報', '品質情報', '均質番号']
    columns_to_include = []
    for idx, cell in enumerate(sheet[6][1:], start=2):  # 6行目のB列以降をチェック
        if not any(excluded_label in str(cell.value) for excluded_label in excluded_labels):
            columns_to_include.append(idx)  # 列のインデックスを保存

    # フィルタリングされたラベルを保持するリスト
    filtered_labels = [labels[idx-2] for idx in columns_to_include]

    # データを[ラベル][年度][月][旬][データ]の形で取得する辞書
    meteorological_data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    years_collected = set()  # 取得した年度を記録
    none_count = 0  # 欠損データカウント

    for row in range(7, sheet.max_row + 1):  # 7行目からデータが始まる
        year_month_day = sheet.cell(row=row, column=1).value
        if not year_month_day:
            continue

        # 年月日から年、月、日を分離して、会計年度を判定
        year, month, day = map(int, year_month_day.split('/'))

        # 会計年度の判断（9月〜翌年8月で年度を切り替える）
        fiscal_year = year if month <= 8 else year + 1

        # 年度フィルタリング（指定された範囲内のみ取得）
        if start_year and fiscal_year < start_year:
            continue
        if end_year and fiscal_year > end_year:
            continue

        years_collected.add(fiscal_year)  # 取得した年度を記録

        # 旬の判定（1〜10日が上旬、11〜20日が中旬、21〜月末が下旬）
        if day <= 10:
            period = '上旬'
        elif 11 <= day <= 20:
            period = '中旬'
        else:
            period = '下旬'

        # 各ラベルのデータを取得
        for col_idx, label in zip(columns_to_include, filtered_labels):
            if month not in meteorological_data[label][fiscal_year]:
                meteorological_data[label][fiscal_year][month] = {}

            # セルの値を取得
            value = sheet.cell(row=row, column=col_idx).value

            # 空白セル（None）の処理：スキップし、スキップ情報を表示
            if value is None:
                none_count += 1
                """ if verbose:
                    print(f"データスキップ: {fiscal_year}年 {month}月 ({label})") """
                continue  # このデータをスキップして次の処理へ    

            # 各月の旬ごとにデータを格納
            meteorological_data[label][fiscal_year][month][period] = value

    # 取得したデータの年度範囲を表示
    if years_collected:
        min_year = min(years_collected)
        max_year = max(years_collected)
        print(f"取得されたデータの年度範囲: {min_year}年 ～ {max_year}年")
    else:
        print("Warning: 指定された範囲に有効なデータがありませんでした。")

    if verbose:
        print(f"labels: {filtered_labels}")
        print(f"総データ数: {sum(len(v) for y in meteorological_data.values() for m in y.values() for v in m.values())}")
        print(f"欠損データ (None) の数: {none_count}")

    print("気象データ取得完了。")
    print("---------------------------------------------------")

    # データを返す
    return meteorological_data

def calculate_correlations_and_export_with_formatting(meteorological_data, output_file, missing_threshold=0.2):
    """
    気象データから相関行列を計算し、欠損データがある変数を削除。
    欠損が多い変数（欠損率が missing_threshold を超えるもの）を削除し、絶対値が0.5以上の相関係数のみをエクセルファイルに出力。
    また、相関係数の絶対値が大きいほど色分けして表示する。
    
    Parameters:
    - meteorological_data: dict 気象データ
    - output_file: str 出力するエクセルファイルのパス
    - missing_threshold: float 欠損値の割合がこの値を超えるカラムを削除する
    """
    # データをフラット化して、年度・月ごとのデータを一つのデータフレームにする
    flattened_data = []
    for label, years_data in meteorological_data.items():
        for year, months_data in years_data.items():
            for month, periods_data in months_data.items():
                # 旬の情報があるかどうかを確認
                if isinstance(periods_data, dict):
                    # 旬の情報がある場合
                    for period, value in periods_data.items():
                        # 新しいラベルを作成 (例: '平均気温_12月_上旬')
                        new_label = f"{label}_{month}月_{period}"
                        flattened_data.append([year, new_label, value])
                else:
                    # 旬の情報がない場合
                    new_label = f"{label}_{month}月"
                    flattened_data.append([year, new_label, periods_data])
    
    # フラット化したデータをデータフレームに変換
    df = pd.DataFrame(flattened_data, columns=['Year', 'Label', 'Value'])

    # データフレームをピボットして、年度ごとに各ラベルを列として整理
    pivot_df = df.pivot_table(index='Year', columns='Label', values='Value')

    # 欠損値が多い変数（missing_threshold以上の欠損がある変数）を削除
    missing_data_percentage = pivot_df.isnull().mean()  # 欠損データの割合を計算
    removed_columns = missing_data_percentage[missing_data_percentage > missing_threshold].index.tolist()  # 削除する変数名を取得
    pivot_df = pivot_df.drop(columns=removed_columns)  # 欠損が多いカラムを削除

    # 削除された変数を表示
    print(f"削除された変数（欠損データが {missing_threshold * 100}% を超えたもの）:")
    for column in removed_columns:
        print(column)

    # 相関行列を計算
    correlation_matrix = pivot_df.corr()

    # 相関行列の絶対値が0.5以上の組み合わせのみを抽出
    filtered_corr_matrix = correlation_matrix[(correlation_matrix.abs() >= 0.5)]

    # エクセルに出力
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 元データを保存
        pivot_df.to_excel(writer, sheet_name='Flattened Data')  
        
        # 相関行列の絶対値が0.5以上のもののみ保存
        filtered_corr_matrix.to_excel(writer, sheet_name='Filtered Correlation Matrix')

        # ワークブックとシートの取得
        workbook = writer.book
        sheet = workbook['Filtered Correlation Matrix']

        # 色分けの設定
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):  # 数値型か確認
                    if abs(cell.value) >= 0.9:
                        fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # 濃い赤
                    elif abs(cell.value) >= 0.8:
                        fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")  # 赤
                    elif abs(cell.value) >= 0.7:
                        fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # オレンジ
                    elif abs(cell.value) >= 0.6:
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
                    elif abs(cell.value) >= 0.5:
                        fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # 薄い黄色
                    else:
                        fill = None  # 色をつけない


                    if fill:
                        cell.fill = fill

    print(f"フィルタリングされた相関行列が {output_file} に保存され、色分けが適用されました。")


if __name__ == '__main__': 
    # 使用例
    
    # ファイルパス
    month_data_path = 'resources/meteorological_data/nandan_month_12-8.xlsx'  
    syun_data_path = 'resources/meteorological_data/nandan_syun_12-8.xlsx'

    start_year = 1980
    end_year = 2020
    # 気象データをインポート
    meteorological_data_month = import_meteorological_month_data(month_data_path, verbose=True)
    meteorological_data_syun = import_meteorological_syun_data(syun_data_path, verbose=True)

    # 相関行列を計算し、エクセルに出力
    # output_file = 'test_correlation_output.xlsx'  # 出力するエクセルファイルパス
    # calculate_correlations_and_export_with_formatting(meteorological_data_month, output_file)

    # データの一部を表示して確認
    """ for label in meteorological_data_month:
        print(f"\n{label}の年度ごとのデータ:")
        for year, months_data in meteorological_data_month[label].items():
            print(f"{year}年度:")
            for month, value in months_data.items():
                if value is not None:
                    print(f"{month}: {value:.2f}")
                else:
                    print(f"{month}: None") 

    # データの表示
    for label, years_data in meteorological_data_syun.items():
        print(f"\n--- {label} ---")
        for year, months_data in years_data.items():
            print(f"{year}年度:")
            for month, periods_data in months_data.items():
                print(f"  {month}月:")
                for period, value in periods_data.items():
                    print(f"    {period}: {value:.2f}")  """