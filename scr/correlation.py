import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
import japanize_matplotlib
import os
import re
from scipy import stats

from scipy.stats import pearsonr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# --- period名の対応マップ（収穫日分析用） ---


# ------------------------------
# 【メイン関数】
# ------------------------------
def analyze_and_export_correlation(df, target_type="cor1", target_col="log_incidence", output_path="results_it"):
    print("---------------------------------------------------")
    print("相関分析を開始します。")
    print(f"target_type: {target_type}")
    print(f"出力先: {output_path}")

    if target_type == "cor1":
        analyze_weather_by_period(df, target_col,f"{output_path}/weather_corr.xlsx")
        export_period_weather_corr_matrix(df, target_col, output_path)

    elif target_type == "cor2":
        analyze_features_by_period(df, target_col, f"{output_path}/feature_corr.xlsx")
        export_period_feature_corr_matrix(df, target_col, output_path)


    else:
        raise ValueError("target_type must be 'cor1', 'cor2', or 'cor3'")

    print(f"相関分析結果を {output_path} に保存しました。")
    print("---------------------------------------------------")

# ------------------------------
# 相関計算と有意性評価の補助関数
# ------------------------------

def calculate_corr_and_pval(x, y):
    """
    2つのSeriesの相関係数、p値、サンプル数（NaN除外後）を返す
    """
    mask = x.notna() & y.notna()
    n = mask.sum()
    if n < 3:
        return np.nan, np.nan, n
    corr, pval = stats.pearsonr(x[mask], y[mask])
    return corr, pval, n

def compute_correlation_and_significance(subset, target_col, valid_cols, period, output_dir, scatter_threshold=0.4):
    os.makedirs(output_dir, exist_ok=True)
    corr_results = []
    pval_results = []
    n_results = []

    for feat in valid_cols:
        corr, pval, n = calculate_corr_and_pval(subset[feat], subset[target_col])
        corr_results.append((feat, corr))
        pval_results.append((feat, pval))
        n_results.append((feat, n))

    corr_df = pd.DataFrame(corr_results, columns=["feature", f"corr_with_{target_col}"])
    pval_df = pd.DataFrame(pval_results, columns=["feature", "p_value"])
    n_df = pd.DataFrame(n_results, columns=["feature", "n_samples"])

    # ランキング
    ranking_df = corr_df.copy()
    ranking_df["abs_corr"] = np.abs(ranking_df[f"corr_with_{target_col}"])
    ranking_df = ranking_df.sort_values("abs_corr", ascending=False).reset_index(drop=True)
    ranking_df = ranking_df.drop(columns=["abs_corr"])
    ranking_df = ranking_df.merge(pval_df, on="feature").merge(n_df, on="feature")

    # Excel出力
    excel_path = os.path.join(output_dir, f"{period}_corr_ranking.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        ranking_df.to_excel(writer, sheet_name="ranking", index=False)
    print(f"  - 相関係数ランキング保存: {excel_path}")

    # 有意な相関のログ出力
    signif_df = ranking_df[ranking_df["p_value"] < 0.05]
    if not signif_df.empty:
        print(f"  - 有意な相関（p<0.05）:")
        for _, row in signif_df.iterrows():
            print(f"    {row['feature']}: p={row['p_value']:.3g}, corr={row[f'corr_with_{target_col}']:.2f}, n={row['n_samples']}")
    else:
        print("  - 有意な相関はありません（p<0.05）")

    # 散布図（強相関のみ）
    strong = ranking_df[np.abs(ranking_df[f"corr_with_{target_col}"]) >= scatter_threshold]
    for _, row in strong.iterrows():
        feat = row['feature']
        plt.figure(figsize=(5, 4))
        sns.regplot(data=subset, x=feat, y=target_col, line_kws={'color': 'red'})
        plt.title(f"{period}: {feat} vs {target_col}\n(corr={row[f'corr_with_{target_col}']:.2f}, n={row['n_samples']})")
        plt.xlabel(feat)
        plt.ylabel(target_col)
        plt.tight_layout()
        img_path = os.path.join(output_dir, f"{period}_{feat}_scatter.pdf")
        plt.savefig(img_path)
        plt.close()
        print(f"    - 強相関: {feat} の散布図保存: {img_path}")




# ------------------------------
# 【1】weather_by_period：periodごとに気象 vs incidence
# ------------------------------

def analyze_weather_by_period(df, target_col, output_path):
    """
    各 period ごとに weather特徴量（50%以上NaNは除外）と target_col の相関係数・p値・サンプル数を出力
    ＋ランキング表・散布図も出力（compute_correlation_and_significanceを呼ぶ）
    """
    weather_cols = [col for col in df.columns if col.endswith("_avg") or col.endswith("_sum")]
    periods = df["period"].dropna().unique()
    output_dir = os.path.dirname(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        any_written = False

        for period in periods:
            subset = df[df["period"] == period].copy()
            # 50%以上NaNは除外
            valid_cols = [col for col in weather_cols if subset[col].notna().sum() / len(subset) >= 0.5]
            removed_cols = [col for col in weather_cols if col not in valid_cols]
            if removed_cols:
                print(f"[{period}] 50%以上NaNで除外されたカラム: {removed_cols}")

            if not valid_cols or subset[valid_cols + [target_col]].dropna().empty:
                print(f"[!] {period} は有効なデータがなくスキップされました")
                continue

            # 相関・p値・n計算
            corr_list, pval_list, n_list = [], [], []
            for feat in valid_cols:
                corr, pval, n = calculate_corr_and_pval(subset[feat], subset[target_col])
                corr_list.append(corr)
                pval_list.append(pval)
                n_list.append(n)
            stats_df = pd.DataFrame({
                "feature": valid_cols,
                "corr": corr_list,
                "p_value": pval_list,
                "n_samples": n_list
            }).set_index("feature")

            # Excel書き出し（このperiodだけの相関詳細表）
            stats_df.to_excel(writer, sheet_name=str(period)[:30])
            any_written = True

            # --------- ここでcompute_correlation_and_significanceを呼び出し -----------
            # これによりperiodごとに「ランキング表」「有意ログ」「散布図」も自動出力
            compute_correlation_and_significance(
                subset, target_col, valid_cols, period,
                output_dir=output_dir
            )
            # ---------------------------------------------------------------------------

        if not any_written:
            pd.DataFrame({"Warning": ["No valid data found."]}).to_excel(writer, sheet_name="empty")


weather_order = [
    "mean_temp",
    "max_temp",
    "min_temp",
    "total_precip",
    "sunshine_duration",
    "mean_wind_speed",
    "max_wind_speed",
    "mean_humidity",
    "mean_vapor_pressure"
]
weather_map = {name: i for i, name in enumerate(weather_order)}
# avg用
day_ranges = ["1-7", "8-14", "15-30", "31-60", "61-90", "91-120", "121-150", "151-180", "181-210", "211-240"]
day_range_map = {rng: i for i, rng in enumerate(day_ranges)}
# sum用
days_priority = [7, 14, 30, 60, 90, 120, 150, 180, 210, 240]
days_map = {str(d): i for i, d in enumerate(days_priority)}

def sort_weather_features_full(index_list, mode="avg"):
    """
    mode: "avg" or "sum"
    """
    def parse_feature(x):
        x = x.strip()
        if mode == "avg":
            # 例: mean_temp_1-7days_avg
            m = re.match(r"([a-zA-Z_]+)_(\d+-\d+)days_avg$", x)
            if m:
                feat = m.group(1)
                rng = m.group(2)
                feat_idx = weather_map.get(feat, 99)
                rng_idx = day_range_map.get(rng, 99)
                return (feat_idx, rng_idx, x)
        elif mode == "sum":
            # 例: mean_temp_7days_sum
            m = re.match(r"([a-zA-Z_]+)_(\d+)days_sum$", x)
            if m:
                feat = m.group(1)
                days = m.group(2)
                feat_idx = weather_map.get(feat, 99)
                days_idx = days_map.get(days, 99)
                return (feat_idx, days_idx, int(days), x)
        # fallback
        return (99, 99, x)
    return sorted(index_list, key=parse_feature)

def export_period_weather_corr_matrix(df, target_col, output_path):
    periods = df["period"].dropna().unique()
    avg_cols = [col for col in df.columns if col.endswith("_avg")]
    sum_cols = [col for col in df.columns if col.endswith("_sum")]
    outdict_corr = {}
    outdict_n = {}

    for suffix, weather_cols in [("_avg", avg_cols), ("_sum", sum_cols)]:
        period_corr_dict = {}
        period_n_dict = {}
        for period in periods:
            subset = df[df["period"] == period].copy()
            valid_cols = [col for col in weather_cols if subset[col].notna().sum() / len(subset) >= 0.5]
            removed_cols = [col for col in weather_cols if col not in valid_cols]
            if removed_cols:
                print(f"[{period}][{suffix}] 50%以上NaNで除外されたカラム: {removed_cols}")

            if not valid_cols or subset[valid_cols + [target_col]].dropna().empty:
                continue

            period_corrs = {}
            period_ns = {}
            for col in valid_cols:
                corr, _, n = calculate_corr_and_pval(subset[col], subset[target_col])
                period_corrs[col] = corr
                period_ns[col] = n
            if period_corrs:
                period_corr_dict[period] = pd.Series(period_corrs)
                period_n_dict[period] = pd.Series(period_ns)

        if period_corr_dict:
            corr_df = pd.DataFrame(period_corr_dict)
            n_df = pd.DataFrame(period_n_dict)

            # 並び替え（avg/sumで切り替え）
            mode = "avg" if suffix == "_avg" else "sum"
            sorted_idx = sort_weather_features_full(corr_df.index, mode=mode)
            corr_df = corr_df.loc[sorted_idx]
            n_df = n_df.loc[sorted_idx]

            outdict_corr[suffix] = corr_df
            outdict_n[suffix] = n_df

    # Excel出力
    with pd.ExcelWriter(f"{output_path}/period_weather_corr.xlsx", engine="openpyxl") as writer:
        for suffix in outdict_corr:
            outdict_corr[suffix].to_excel(writer, sheet_name=f"Weather{suffix}")
            outdict_n[suffix].to_excel(writer, sheet_name=f"n_samples{suffix}")
    print("縦=weather, 横=period の相関行列（corr, n）をエクセル出力しました。")



# ------------------------------
# 【2】harvest：収穫日 vs 過去の発病率（ピボットで整理）
# ------------------------------
def analyze_features_by_period(df, target_col, output_path):
    """
    periodごとに、気象列・メタ情報以外の特徴量とtarget_colの相関・p値・サンプル数を出力
    """
    meta_cols = [
        "brand", "year", "period", "date",
        "incidence", "log_incidence", "days_from_jan1"
    ]
    weather_cols = [col for col in df.columns if col.endswith("_avg") or col.endswith("_sum")]
    candidate_cols = [
        col for col in df.columns
        if col not in meta_cols and col not in weather_cols and col != target_col
    ]
    periods = df["period"].dropna().unique()
    output_dir = os.path.dirname(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        any_written = False

        for period in periods:
            print("---------------------------------------------------")
            print(f"Processing period: {period}")
            subset = df[df["period"] == period].copy()
            # 50%以上NaNで除外
            valid_cols = [col for col in candidate_cols if subset[col].notna().sum() / len(subset) >= 0.5]
            removed_cols = [col for col in candidate_cols if col not in valid_cols]
            if removed_cols:
                print(f"number of valid columns: {len(valid_cols)}")
                # print(f"[{period}] 50%以上NaNで除外されたカラム: {removed_cols}")

            if not valid_cols or subset[valid_cols + [target_col]].dropna().empty:
                print(f"[!] {period} は有効なデータがなくスキップされました")
                continue

            # 相関・p値・n計算
            corr_list, pval_list, n_list = [], [], []
            for feat in valid_cols:
                corr, pval, n = calculate_corr_and_pval(subset[feat], subset[target_col])
                corr_list.append(corr)
                pval_list.append(pval)
                n_list.append(n)
            stats_df = pd.DataFrame({
                "feature": valid_cols,
                "corr": corr_list,
                "p_value": pval_list,
                "n_samples": n_list
            }).set_index("feature")

            # Excel書き出し
            stats_df.to_excel(writer, sheet_name=str(period)[:30])
            any_written = True

            # 有意性評価・散布図出力などは共通関数
            compute_correlation_and_significance(subset, target_col, valid_cols, period, output_dir)

        if not any_written:
            pd.DataFrame({"Warning": ["No valid data found."]}).to_excel(writer, sheet_name="empty")

def export_period_feature_corr_matrix(df, target_col, output_path):
    """
    縦軸：各種特徴量（気象特徴量・メタ情報以外）
    横軸：period
    各セル：periodごとの特徴量と target_col の相関係数・サンプル数
    """
    meta_cols = [
        "brand", "year", "period", "date",
        "incidence", "log_incidence", "days_from_jan1"
    ]
    weather_cols = [col for col in df.columns if col.endswith("_avg") or col.endswith("_sum")]
    candidate_cols = [
        col for col in df.columns
        if col not in meta_cols and col not in weather_cols and col != target_col
    ]
    periods = df["period"].dropna().unique()
    corr_data = {}
    n_data = {}

    for period in periods:
        subset = df[df["period"] == period].copy()
        # 50%以上NaNで除外
        valid_cols = [col for col in candidate_cols if subset[col].notna().sum() / len(subset) >= 0.5]
        removed_cols = [col for col in candidate_cols if col not in valid_cols]
        if removed_cols:
            print(f"[{period}] 50%以上NaNで除外されたカラム: {removed_cols}")

        if not valid_cols or subset[valid_cols + [target_col]].dropna().empty:
            continue

        period_corrs = {}
        period_ns = {}
        for col in valid_cols:
            corr, _, n = calculate_corr_and_pval(subset[col], subset[target_col])
            period_corrs[col] = corr
            period_ns[col] = n
        if period_corrs:
            corr_data[period] = pd.Series(period_corrs)
            n_data[period] = pd.Series(period_ns)

    if corr_data:
        corr_df = pd.DataFrame(corr_data)
        n_df = pd.DataFrame(n_data)
        with pd.ExcelWriter(f"{output_path}/period_feature_corr.xlsx", engine="openpyxl") as writer:
            corr_df.to_excel(writer, sheet_name="Features")
            n_df.to_excel(writer, sheet_name="n_samples")
        print("縦=feature, 横=period の相関行列（corr, n）をエクセル出力しました。")
    else:
        print("有効なデータがありませんでした。")


    


