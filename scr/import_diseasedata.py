import pandas as pd
from collections import defaultdict
import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib
from matplotlib.dates import DateFormatter
import seaborn as sns
import os
from openpyxl import load_workbook
import numpy as np
import os
import datetime

def import_disease_data(file_path, brand, start_year, end_year):
    print("---------------------------------------------------")
    print(f"データ読み込み完了: {file_path}")
    print(f"品種: {brand}")
    print(f"年度範囲: {start_year}年 ～ {end_year}年")

    wb = load_workbook(file_path, data_only=True)
    data = []

    # C〜L列 (index 2〜11) に対応する時期ラベル
    period_order = {
        2: "storage_survey",
        3: "harvest",
        4: "late_May",
        5: "early_May",
        6: "late_April",
        7: "early_April",
        8: "late_March",
        9: "early_March",
        10: "late_February",
        11: "early_February"
    }

    for sheet_name in wb.sheetnames:
        # 年度として扱えるシートのみを対象
        try:
            year = int(sheet_name)
        except ValueError:
            print(f"Warning: '{sheet_name}' は年度として扱えません。スキップします。")
            continue

        if not (start_year <= year <= end_year):
            continue

        # 該当シートの読み込み（ヘッダーなし）
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # 品種名と"平均"を探す
        for idx in range(len(df)):
            if str(df.iloc[idx, 0]) == "平均" and str(df.iloc[idx, 1]) == brand:
                target_row = idx
                break
        else:
            print(f"Warning: 品種 '{brand}' が年度 '{sheet_name}' に見つかりません。")
            continue  # 品種が存在しない場合スキップ

        # データ抽出
        for col in range(2, df.shape[1]):
            value = df.iat[target_row, col]

            try:
                if col == 3:
                    date_raw = df.iat[target_row, 12]  # M列（収穫日）
                else:
                    date_raw = df.iat[0, col]  # 通常は1行目の観測日

                date_parsed = pd.to_datetime(date_raw).date()
                days_from_jan1 = (date_parsed - datetime.date(date_parsed.year, 1, 1)).days

            except Exception as e:
                if str(date_raw) not in {"収穫日", "整合性"}:
                    print(f"Warning: {period_label} {year}年 {brand} 品種で日付のパースまたは日数計算に失敗: {date_raw} → {e}")
                    date_parsed = None
                    days_from_jan1 = None
                else:
                    continue

            # 時期名を列インデックスから取得（ない場合は "不明"）
            period_label = period_order.get(col, "unknown")


            if pd.isna(value):
                print(f"Warning: NaN value found in {period_label} {year}年 {brand} 品種")
                value = None

            elif isinstance(value, (int, float)):
                value = float(value)
            else:   
                print(f"Warning: 数値以外の値 '{value}' {period_label} {year}年 {brand} 品種が見つかりました。スキップします。")
                value = None

            data.append({
                "brand": brand,
                "year": year,
                "period": period_label,
                "date": date_parsed,
                 "incidence": value,
                "days_from_jan1": days_from_jan1
            })


    # 保存処理
    os.makedirs("results_it", exist_ok=True)
    output_path = os.path.join("results_it", f"{brand}_disease_data.xlsx")
    save_records_to_excel(data, output_path)

    print("病害データ取得完了。")
    print("---------------------------------------------------")

    return data

def add_incidence_cumsum(disease_df, period_order):
    import numpy as np
    disease_df = disease_df.copy()
    disease_df["incidence_cumsum"] = np.nan

    # periodを順序付きカテゴリ型
    disease_df["period"] = pd.Categorical(disease_df["period"], categories=period_order, ordered=True)

    for (brand, year), g in disease_df.groupby(["brand", "year"]):
        # period順にソート
        g_sorted = g.sort_values("period")
        values = g_sorted["incidence"].values.astype(float)
        idxs = g_sorted.index

        # 直前までの累積和（shiftしてcumsum, 先頭はNone）
        cumsum_shifted = np.insert(np.cumsum(np.nan_to_num(values))[:-1], 0, np.nan)
        # 必要に応じてNaN（観測なし）はそのままNaN
        for i, idx in enumerate(idxs):
            if not np.isnan(values[i]):
                disease_df.at[idx, "incidence_cumsum"] = cumsum_shifted[i]
            else:
                disease_df.at[idx, "incidence_cumsum"] = np.nan

    return disease_df


    # 3. period間のすべての差分を個別カラムで追加
def add_all_incidence_diffs(disease_df, period_order):
    disease_df = disease_df.copy()
    all_diff_cols = []

    for (brand, year), g in disease_df.groupby(["brand", "year"]):
        g_sorted = g.set_index("period").reindex(period_order)
        incidences = g_sorted["incidence"].astype(float).values

        for idx, period in enumerate(period_order):
            for j in range(1, idx + 1):
                prev_p = period_order[j - 1]
                curr_p = period_order[j]
                diff_val = None
                if (
                    j < len(incidences)
                    and not pd.isna(incidences[j])
                    and not pd.isna(incidences[j - 1])
                ):
                    diff_val = incidences[j] - incidences[j - 1]
                col_name = f"diff_{prev_p}_{curr_p}"
                disease_df.loc[
                    (disease_df["brand"] == brand) &
                    (disease_df["year"] == year) &
                    (disease_df["period"] == period), col_name
                ] = diff_val
                if col_name not in all_diff_cols:
                    all_diff_cols.append(col_name)

    return disease_df

def add_past_period_log_actuals(disease_df, period_order, log_col="log_incidence"):
    """
    各period行に、それ以前のperiodのlog変換 incidence（log_incidence）を個別カラムで追加

    Parameters
    ----------
    disease_df : pd.DataFrame
    period_order : list[str]
    log_col : str
        log変換したincidenceのカラム名（例: 'log_incidence'）

    Returns
    -------
    pd.DataFrame
        新カラム（periodごとのlog_incidence）が付与されたDataFrame
    """
    if log_col not in disease_df.columns:
        raise ValueError(f"'{log_col}' カラムがDataFrameに存在しません")

    disease_df = disease_df.copy()
    for (brand, year), g in disease_df.groupby(["brand", "year"]):
        g_sorted = g.set_index("period").reindex(period_order)
        log_incidences = g_sorted[log_col].astype(float)
        for idx, period in enumerate(period_order):
            for j in range(idx):
                prev_p = period_order[j]
                col_name = f"{prev_p}_{log_col}"
                val = log_incidences[prev_p] if not pd.isna(log_incidences[prev_p]) else None
                disease_df.loc[
                    (disease_df["brand"] == brand) &
                    (disease_df["year"] == year) &
                    (disease_df["period"] == period), col_name
                ] = val
    return disease_df




def save_records_to_excel(records, output_path):
    if not records:
        print("⚠️ データが取得されていません。出力はスキップされました。")
        return

    df = pd.DataFrame(records)
    # Excelとして保存
    df.to_excel(output_path, index=False)
    print(f"データを {output_path} に保存しました（{len(df)} 件）。")


def plot_disease_trends_by_year(data, target_name="ターザン", output_dir=None, show=True):
    """
    Plot disease incidence trends by year for a specific variety, using MM-DD observation dates.

    Parameters
    ----------
    data : list of dict
        Output from import_disease_data(), containing keys like brand, year, date, period, incidence.
    target_name : str
        Target variety name (default: "ターザン").
    output_dir : str or None
        Directory to save the plot. If None, the plot is not saved.
    show : bool
        If True, the plot is displayed (default: True).
    """

    # リスト→DataFrame
    df = pd.DataFrame(data)

    # 品種でフィルタ
    df = df[df["brand"] == target_name].copy()
    if df.empty:
        print(f"No data found for variety '{target_name}'.")
        return

    # incidence 欠損を除外
    df = df[df["incidence"].notna()]

    # date列を 2000年のダミー日付に変換（x軸用）
    df['date_dt'] = pd.to_datetime(df['date'].apply(lambda d: f"2000-{d.month:02d}-{d.day:02d}"))

    # プロット
    plt.figure(figsize=(10, 10))
    sns.lineplot(data=df, x='date_dt', y='incidence', hue='year', marker='o', palette='tab20')

    plt.title(f"Disease incidence trends of {target_name}", fontsize=14)
    plt.xlabel("Observation date (MM-DD)")
    plt.ylabel("Incidence rate")

    plt.gca().xaxis.set_major_formatter(DateFormatter("%m-%d"))
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.legend(title="Year", bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.tight_layout()

    # 保存
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{target_name}_incidence_trends.png")
        plt.savefig(output_path, dpi=300)
        print(f"Plot saved to: {output_path}")

    if show:
        plt.show()
    else:
        plt.close()

if __name__ == '__main__':


    file_path = "resources/disease_data/onion_disease_sammary.xlsx"
    output_dir = "results_it"
    os.makedirs(output_dir, exist_ok=True)
    target_name = "ターザン"

    # 病害データをインポート
    disease_data= import_disease_data(file_path, target_name, start_year=1994, end_year=2022)
    plot_disease_trends_by_year(disease_data, target_name=target_name, output_dir=output_dir, show=True)
