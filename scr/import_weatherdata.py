import pandas as pd
from openpyxl import load_workbook
import os
from datetime import timedelta

def save_to_excel(df, output_path):
    """Convert weather data to wide format and save as Excel"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # ✅ Corrected pivot table with English column names
    wide_df = df.pivot_table(index=["year", "date"], columns="weather_item", values="value", aggfunc="first").reset_index()

    wide_df.columns.name = None  # flatten MultiIndex columns

    wide_df.to_excel(output_path, index=False)
    print(f"✅ Weather data saved to: {output_path}")


def extract_meteorological_data(file_path, start_year, end_year, lang="jp"):
    print("---------------------------------------------------")
    print(f"気象データを読み込み中: {file_path}")
    print(f"対象年度: {start_year}年 ～ {end_year}年")
    print(f"出力言語: {'日本語' if lang == 'jp' else '英語'}")

    # 日本語⇔英語変換辞書
    item_jp2en = {
        "平均気温(℃)": "mean_temp",
        "最高気温(℃)": "max_temp",
        "最低気温(℃)": "min_temp",
        "平均湿度(％)": "mean_humidity",
        "平均蒸気圧(hPa)": "mean_vapor_pressure",
        "平均風速(m/s)": "mean_wind_speed",
        "最大風速(m/s)": "max_wind_speed",
        "日照時間(時間)": "sunshine_duration",
        "降水量の合計(mm)": "total_precip"
    }
    # 英語→日本語（逆変換用）
    item_en2jp = {v: k for k, v in item_jp2en.items()}

    wb = load_workbook(file_path, data_only=True)
    records = []

    # ① シートとその年度範囲のマッピングを作成
    sheet_ranges = []
    for sheet_name in wb.sheetnames:
        try:
            sheet_start, sheet_end = map(int, sheet_name.split("-"))
            sheet_ranges.append((sheet_name, sheet_start, sheet_end))
        except:
            print(f"⚠️ シート名 '{sheet_name}' は 'xxxx-yyyy' 形式ではありません。スキップ。")

    # ② 対象年度ごとに処理
    for target_year in range(start_year, end_year + 1):
        matched_sheet = None
        for sheet_name, range_start, range_end in sheet_ranges:
            if range_start <= target_year <= range_end:
                matched_sheet = sheet_name
                break

        if matched_sheet is None:
            print(f"⚠️ 年度 {target_year} に対応するシートが見つかりません。スキップ。")
            continue

        ws = wb[matched_sheet]
        item_names = [cell.value for cell in ws[2]][1:]  # B列以降

        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            raw_date = row[0]
            if raw_date is None:
                continue

            try:
                date = pd.to_datetime(raw_date).date()
            except:
                continue

            year = date.year + 1 if date.month == 12 else date.year
            if year != target_year:
                continue

            for i, value in enumerate(row[1:], start=0):
                jp_item = item_names[i]

                if lang == "en":
                    item = item_jp2en.get(jp_item, jp_item)
                elif lang == "jp":
                    item = jp_item
                else:
                    raise ValueError("lang must be 'jp' or 'en'")

                records.append({
                    "weather_item": item,
                    "year": year,
                    "date": date,
                    "value": value
                })
            count += 1

        print(f"✔ 年度 {target_year} のデータをシート '{matched_sheet}' から {count} 行取得")

    df = pd.DataFrame(records)
    print("気象データ取得完了。")
    print("---------------------------------------------------")
    return df



def get_weather_window_avg(df, year, obs_date, start_day, end_day):
    """
    Compute the average weather values over a given period before the observation date.

    Parameters
    ----------
    df : pd.DataFrame
        Long-format weather data with columns: weather_item, year, date, value
    year : int
        Target year (e.g., 2017)
    obs_date : datetime.date
        Observation date
    start_day : int
        Start of the window (days before obs_date, inclusive)
    end_day : int
        End of the window (days before obs_date, inclusive)

    Returns
    -------
    pd.DataFrame
        Aggregated average per weather item:
        [weather_item, year, period, value]
    """
    start_date = obs_date - timedelta(days=end_day)
    end_date = obs_date - timedelta(days=start_day)

    subset = df[(df["year"] == year) & 
                (df["date"] >= start_date) & 
                (df["date"] <= end_date)].copy()

    result = subset.groupby("weather_item", as_index=False)["value"].mean()

    result["year"] = year
    result["period"] = f"{start_day}-{end_day}days"
    result = result[["weather_item", "year", "period", "value"]]

    return result


def get_multiple_weather_period(df, year, obs_date, period):
    """
    Get weather averages over dynamically selected time windows based on the disease observation period.

    Parameters
    ----------
    df : pd.DataFrame
        Long-format weather data (columns: weather_item, year, date, value)
    year : int
        Year corresponding to disease observation
    obs_date : datetime.date
        Observation date
    period : str
        Disease observation period (e.g., "2月上旬", "収穫日", "貯蔵調査")

    Returns
    -------
    dict of pd.DataFrame
        Dictionary of averaged weather data keyed by day range (e.g., '7days', '30days', ...)
    """
    period_day_mapping = {
        "early_February": [7, 14, 30, 60],
        "late_February": [7, 14, 30, 60],
        "early_March": [7, 14, 30, 60, 90],
        "late_March": [7, 14, 30, 60, 90],
        "early_April": [7, 14, 30, 60, 90, 120],
        "late_April": [7, 14, 30, 60, 90, 120],
        "early_May": [7, 14, 30, 60, 90, 120, 150],
        "late_May": [7, 14, 30, 60, 90, 120, 150],
        "harvest":  [7, 14, 30, 60, 90, 120, 150],
        "storage_survey": [7, 14, 30, 60, 90, 120, 150, 180, 210, 240]
    }


    day_list = period_day_mapping.get(period, [7, 14, 30])

    result = {}
    for i, days in enumerate(day_list):
        if days == 7:
            start_day = 1
            end_day = 7
        else:
            prev = day_list[i - 1] if i > 0 else 7
            start_day = prev + 1
            end_day = days

        df_window = get_weather_window_avg(df, year, obs_date, start_day=start_day, end_day=end_day)
        result[f"{start_day}-{end_day}days_avg"] = df_window

    return result

def get_weather_window_sum(df, year, obs_date, window_days):
    """
    Compute the sum of weather values over the N days before the observation date.
    window_days: 7なら「観測日-7日から観測日-1日まで」。
    """
    start_date = obs_date - timedelta(days=window_days)
    end_date = obs_date - timedelta(days=1)  # 観測日直前まで

    subset = df[
        (df["year"] == year) & 
        (df["date"] >= start_date) & 
        (df["date"] <= end_date)
    ].copy()

    result = subset.groupby("weather_item", as_index=False)["value"].sum()
    result["year"] = year
    result["period"] = f"{window_days}days_sum"
    result = result[["weather_item", "year", "period", "value"]]

    return result

def get_multiple_weather_sum_period(df, year, obs_date, period):
    """
    Get weather sum (accumulation) over N days up to the observation date for each window.
    Returns dict of DataFrames, keyed by day range.
    """
    period_day_mapping = {
        "early_February": [7, 14, 30, 60],
        "late_February": [7, 14, 30, 60],
        "early_March": [7, 14, 30, 60, 90],
        "late_March": [7, 14, 30, 60, 90],
        "early_April": [7, 14, 30, 60, 90, 120],
        "late_April": [7, 14, 30, 60, 90, 120],
        "early_May": [7, 14, 30, 60, 90, 120, 150],
        "late_May": [7, 14, 30, 60, 90, 120, 150],
        "harvest":  [7, 14, 30, 60, 90, 120, 150],
        "storage_survey": [7, 14, 30, 60, 90, 120, 150, 180, 210, 240]
    }

    day_list = period_day_mapping.get(period, [7, 14, 30])

    result = {}
    for days in day_list:
        df_window = get_weather_window_sum(df, year, obs_date, window_days=days)
        result[f"{days}days_sum"] = df_window

    return result



if __name__ == "__main__":
    file_path = "resources/meteorological_data/1990_2025_sumoto.xlsx"
    df = extract_meteorological_data(file_path, 1998, 2002)