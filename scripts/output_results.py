import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import japanize_matplotlib
import seaborn as sns
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import re

def sanitize_filename(s):
    # ファイル名に使えない文字を `_` に置き換える
    return re.sub(r'[\\/:"*?<>|]+', '_', s)

def save_results_to_excel(models, predictors_list, filename="regression_results.xlsx", folder="outputs"):
    """
    重回帰分析の結果をExcelファイルに出力する関数（単体/リスト対応、標準化回帰係数追加）。

    Args:
        models (list): statsmodelsのモデルオブジェクトのリスト
        predictors_list (list): 説明変数のリストのリスト
        filename (str): 出力ファイル名
        folder (str): 出力フォルダ名（デフォルト: "outputs"）

    Returns:
        None
    """
    print("---------------------------------------------------")
    print("回帰分析の結果をExcelファイルに保存します。")

    # 出力フォルダの作成
    os.makedirs(folder, exist_ok=True)
    
    if not isinstance(models, list):
        models, predictors_list = [models], [predictors_list]

    results = []
    for i, (model, predictors) in enumerate(zip(models, predictors_list)):
        # モデル名を作成（例: Best1, Best2, ...）
        model_name = f"Best{i+1}"
        model_header = pd.DataFrame([[model_name]], columns=[""])  # 1行だけのDataFrame
        
        # モデルのサマリーを取得
        summary_info_df = pd.DataFrame(model.summary2().tables[0])

        # 回帰係数テーブルを取得
        summary_df = pd.DataFrame(model.summary2().tables[1])

        # 標準化回帰係数（Standardized Beta）の計算    
        std_X = model.model.exog[:, 1:].std(axis=0)
        std_Y = model.model.endog.std()
        standardized_betas = model.params[1:] * (std_X / std_Y)

        # テーブルの整形
        summary_df.columns = ["Coef.", "Std Err", "t", "P>|t|", "[0.025", "0.975]"]
        summary_df.insert(0, "Variable", model.params.index)
        summary_df["Standardized Beta"] = pd.Series(standardized_betas, index=model.params.index[1:])
        empty_row = pd.DataFrame([[""] * len(summary_df.columns)], columns=summary_df.columns)

        # 各モデルのデータを整理
        model_result = pd.concat([model_header, summary_info_df, summary_df, empty_row])
        results.append(model_result)

    # すべての結果を結合
    results_df = pd.concat(results, ignore_index=True)

    # Excelに保存
    file_path = os.path.join(folder, filename)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, index=False, sheet_name="Regression Results")

    print(f"Results saved to {file_path}")
    print("---------------------------------------------------")


def plot_multiple_regression(models, final_df, filename_prefix="mulitiple_regression_plot", folder="outputs", show_years=False):
    """
    重回帰分析の結果（Fitted vs Actual）をプロットして画像ファイルとして保存する関数（複数のモデル対応）。
    
    Args:
        models (list): statsmodelsのモデルオブジェクトのリスト
        final_df (pd.DataFrame): 回帰分析の結果を含むデータフレーム（'発病率（実測値）', '発病率（予測値）', '年度', 'モデルID' を含む）
        filename_prefix (str): 画像ファイルのプレフィックス
        folder (str): 出力フォルダ名（デフォルト: "outputs"）
        show_years (bool): 各点に年度を表示するか（デフォルト: False）

    Returns:
        None
    """
    print("---------------------------------------------------")
    print("重回帰分析の結果（Fitted vs Actual）をプロットして保存します。")

    # 出力フォルダの作成
    os.makedirs(folder, exist_ok=True)

    # モデルごとにプロットを作成
    for i, (model_id, df) in enumerate(final_df.groupby("モデルID")):
        y = df["発病率（実測値）"]
        predictions = df["発病率（予測値）"]
        years = df["年度"] if "年度" in df.columns else None  # 年度が存在する場合のみ取得

        plt.figure(figsize=(10, 6))
        plt.scatter(y, predictions, alpha=0.6, edgecolors='k', label="Fitted vs Actual")
        plt.plot([y.min(), y.max()], [y.min(), y.max()], 'k--', lw=2, label="Ideal Fit")

        # 年度をプロット上に表示するオプション
        if show_years and years is not None:
            for actual, predicted, year in zip(y.to_numpy(), predictions, years):
                plt.annotate(str(year), (actual, predicted), fontsize=6, ha='right', va='bottom', alpha=1.0, color="black")

        # 軸とタイトルの設定
        plt.xlabel('Actual Values')
        plt.ylabel('Fitted Values')
        plt.title(f"multiple Regression Results - {model_id}")

        # models[i]` から R² と Adj R² を取得
        r_squared = models[i].rsquared
        adj_r_squared = models[i].rsquared_adj

        plt.text(0.1, 0.9, f'R-squared: {r_squared:.2f}', transform=plt.gca().transAxes)
        plt.text(0.1, 0.85, f'Adj R-squared: {adj_r_squared:.2f}', transform=plt.gca().transAxes)

        plt.legend()
        plt.grid(True)

        # 画像を保存
        filename = f"{folder}/{filename_prefix}_{model_id}.pdf"
        plt.savefig(filename, format="pdf")
        plt.close()
        print(f"Plot saved as {filename}")

    print("---------------------------------------------------")


def plot_explanatory_vs_target(df_final, filename_prefix="explanatory_vs_target", folder="outputs"):
    """
    重回帰分析の結果を可視化する関数（相関係数をプロット）
    各説明変数と目的変数をプロットして画像ファイルとして保存する関数。

    Args:
        df_final (pd.DataFrame): モデルごとの予測結果を含むデータフレーム（発病率、選択された説明変数、年度を含む）
        filename_prefix (str): 画像ファイルのプレフィックス
        folder (str): 出力フォルダ名（デフォルト: "outputs"）

    Returns:
        None
    """
    print("---------------------------------------------------")
    print("重回帰分析の結果（各説明変数と目的変数）をプロットします。")

    # 出力フォルダの作成
    os.makedirs(folder, exist_ok=True)

    # モデルごとにプロットを作成
    for model_id, df in df_final.groupby("モデルID"):
        y_actual = df["発病率（実測値）"]  
        years = df["年度"]  

        # 選択された説明変数を取得
        selected_features = [col for col in df.columns if col not in ["品種", "年度", "場所", "モデルID", "発病率（実測値）", "発病率（予測値）", "誤差"]]

        for feature in selected_features:
            x = df[feature]

            # `x` の値が全て同じか、NaN / Inf を含む場合はスキップ
            if np.isnan(x).any() or np.isinf(x).any():
                print(f"Skipping {feature} in {model_id}: Contains NaN or Inf.")
                continue  
            elif np.all(x == x.iloc[0]): 
                print(f"Skipping {feature} in {model_id}: All values are the same.")
                continue  #

            plt.figure(figsize=(8, 6))
            sns.scatterplot(x=x, y=y_actual, color="blue", label="Actual Values")  # 実測値

            # 各データ点の横に年度を表示
            for actual, feature_value, year in zip(y_actual, x, years):
                plt.annotate(str(year), (feature_value, actual), fontsize=8, ha='right', alpha=0.7)

            # 回帰直線をプロット
            coef = np.polyfit(x, y_actual, 1)  # 1次回帰（直線）
            regression_line = np.poly1d(coef)
            x_range = np.linspace(x.min(), x.max(), 100)
            plt.plot(x_range, regression_line(x_range), 'k--', lw=2, label="Regression Line")

            # 相関係数を計算
            correlation = np.corrcoef(x, y_actual)[0, 1]

            # 軸とタイトルの設定
            plt.xlabel(feature)
            plt.ylabel("発病率")
            plt.title(f"{feature}_vs_発病率 - {model_id}")

            # 相関係数をプロットに追加
            plt.text(0.1, 0.9, f'Correlation: {correlation:.2f}', transform=plt.gca().transAxes)

            plt.legend()
            plt.grid(True)


            # 安全なファイル名に変換して保存
            safe_feature = sanitize_filename(feature)
            filename = f"{folder}/{safe_feature}_vs_発病率_{model_id}.pdf"
            plt.savefig(filename, format="pdf")
            plt.close()
            print(f"Plot saved as {filename}")

    print("---------------------------------------------------")


def save_correlation_results_with_colors(correlation_df, filename):
    """
    相関係数の値に応じて細かく色を付けてエクセルに保存する。

    :param correlation_df: 相関係数を含むデータフレーム
    :param filename: 保存するエクセルファイルの名前
    """

    # Workbookを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "Correlation Analysis"

    # ヘッダーを書き込み
    ws.append(correlation_df.columns.tolist())

    # 色の定義（6段階）
    fill_very_high_positive = PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # 濃い緑（非常に強い正の相関）
    fill_high_positive = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 緑（強い正の相関）
    fill_medium_positive = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 薄い緑（中程度の正の相関）
    fill_weak = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 黄色（弱い相関）
    fill_medium_negative = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 薄い赤（中程度の負の相関）
    fill_high_negative = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")  # 赤（強い負の相関）
    fill_very_high_negative = PatternFill(start_color="800000", end_color="800000", fill_type="solid")  # 濃い赤（非常に強い負の相関）

    # データを書き込み & 条件付き書式を適用
    for row in correlation_df.itertuples(index=False):
        気象情報, 相関係数 = row
        ws.append([気象情報, 相関係数])

        # 色を設定（細かい条件分岐）
        cell = ws[f"B{ws.max_row}"]
        if 相関係数 >= 0.5:
            cell.fill = fill_very_high_positive  # 濃い緑（非常に強い正の相関）
        elif 0.3 <= 相関係数 < 0.5:
            cell.fill = fill_high_positive  # 緑（強い正の相関）
        elif 0.2 <= 相関係数 < 0.3:
            cell.fill = fill_medium_positive  # 薄い緑（中程度の正の相関）
        elif -0.2 <= 相関係数 < 0.2:
            cell.fill = fill_weak  # 黄色（弱い相関）
        elif -0.3 <= 相関係数 < -0.2:
            cell.fill = fill_medium_negative  # 薄い赤（中程度の負の相関）
        elif -0.5 <= 相関係数 < -0.3:
            cell.fill = fill_high_negative  # 赤（強い負の相関）
        else:
            cell.fill = fill_very_high_negative  # 濃い赤（非常に強い負の相関）

    # Excelに保存
    wb.save(filename)
